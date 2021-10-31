import psycopg2
import psycopg2.extras
import os
import xlsxwriter
from openpyxl import load_workbook


def make_dict(result, columns):
    """ making dictionary for result """
    results = []
    for row in result:
        row_dict = {}
        for i, col in enumerate(columns):
            row_dict[col.name] = row[i]
        results.append(row_dict)
    return results


def excel_creator(result):
    for item in result:
        get_next_dict = iter([item])
        headers = item.keys()
        # create excel file if it does not exist
        if not os.path.isfile('example.xlsx'):
            book = xlsxwriter.Workbook('example.xlsx')
            sheet = book.add_worksheet('example')
            for (idx, header) in enumerate(headers):
                sheet.write(0, idx, header)
            book.close()
        # open the files and start the loop
        with open('example.xlsx', 'a+') as csv_file:
            book = load_workbook('example.xlsx')
            sheet = book.get_sheet_by_name('example')

            # loop through all dictionaries
            for d in get_next_dict:
                values = [d[key] for key in headers]
                # write to excel file
                sheet.append(values)
            book.save(filename='example.xlsx')


class Postgres:
    def __init__(self, password, host, port, database=None):
        """
        initiate to postgresql connector for creating db, adding table, insert or selecting data ...

        :param password: pass of the db.
        :param host: host of the db server.
        :param port: port of the db server.
        :param database: name of database.
        """
        self.password = password
        self.host = host
        self.port = port
        """ create a new connection to postgres """
        self.connection = psycopg2.connect(
            database=database, user='postgres', password=self.password, host=self.host, port=self.port
        )
        self.connection.autocommit = True

        """ Creating a cursor object using the cursor() method """
        self.cursor = self.connection.cursor(cursor_factory=psycopg2.extras.DictCursor)

        """ Executing an MYSQL function using the execute() method """
        self.cursor.execute("select version()")
        data = self.cursor.fetchone()
        print("Connection established to: ", (database, data))

    def create_db(self, db_name):
        """
        create a Database in postgresql.
        :param db_name: name of the db
        :return:
        """
        self.cursor.execute("CREATE database %s" % db_name)

    def connect_db(self, db_name):
        """
        connect to a specific database.
        :param db_name:
        :return:
        """
        if db_name is not None:
            self.cursor.close()
            self.connection.close()
            self.__init__(self.password, self.host, self.port, db_name)

    def create_table(self, table_name, query):
        """
        create table into the database
        :param query: define columns of table
        FIRST_NAME CHAR(20) NOT NULL,
        LAST_NAME CHAR(20),
        AGE INT,
        SEX CHAR(1),
        INCOME FLOAT
        :param table_name: define table name.
        :return:
        """

        """ Doping table if already exists. """
        table_name = table_name.upper()
        self.cursor.execute("DROP TABLE IF EXISTS %s CASCADE" % table_name)

        """ Creating table as per requirement """
        self.cursor.execute("CREATE TABLE %s(%s)" % (table_name, query))

    def insert_data(self, table_name, colmuns, values):
        """
        insert data to a specific table of database.
        :param table_name: name of the table which is going to insert data.
        :param query:
        :return:
        """
        self.cursor.execute("INSERT INTO %s (%s) VALUES %s" % (table_name, colmuns, values))

    def select(self, column, table_name):
        """
        select data from a specific table of database.
        :param column: if u want get everything put *
        :param table_name: name of the table which is going to insert data.
        :return:
        """

        """ executing the query """
        self.cursor.execute("SELECT %s from %s" % (column, table_name))
        columns = list(self.cursor.description)
        result = self.cursor.fetchall()
        return make_dict(result, columns)

    def select_where(self, column, table_name, where, order=None):
        """
        select data from a specific table of database. with where clause.
        :param order: sorting order
        :param where: where clause
        :param column: if u want get everything put *
        :param table_name: name of the table which is going to insert data.
        :return:
        """
        if order is not None:
            self.cursor.execute("SELECT %s from %s WHERE %s ORDER BY %s" % (column, table_name, where, order))
            columns = list(self.cursor.description)
            result = self.cursor.fetchall()
            return make_dict(result, columns)
        else:
            self.cursor.execute("SELECT %s from %s WHERE %s" % (column, table_name, where))
            columns = list(self.cursor.description)
            result = self.cursor.fetchall()
            return make_dict(result, columns)

    def report(self, column, table_name, order=None, where=None):
        """
        report database entities in excel files for user.

        :param column: which column need to be in the excel file
        :param table_name: name of the targeted table
        :param order: ordering data for example in order of date
        :param where: where clause for specific data
        :return:
        """
        try:
            if where is not None:
                report_dict = self.select_where(column, table_name, where, order)
                if report_dict is not None:
                    print(report_dict)
            else:
                report_dict = self.select(column, table_name)
                if report_dict is not None:
                    excel_creator(report_dict)

        except Exception as ex:
            print(ex)

    def is_table(self, table_name):
        """
        determine if a table exists
        :param table_name: name of the table
        :return: bool
        """
        self.cursor.execute("select * from information_schema.tables where table_name=%s", (table_name,))
        return bool(self.cursor.rowcount)

    def is_database(self, database_name):
        """
        determine if a database exists.
        :return: bool
        """
        self.cursor.execute("SELECT datname FROM pg_database")
        list_database = self.cursor.fetchall()
        print(list_database)
        if [database_name] in list_database:
            return True
        else:
            return False

    def query(self, query):
        return self.cursor.execute(query)

    def get_cursor(self):
        """
        :return: cursor object.
        """
        return self.cursor

    def disconnect_db(self):
        """
        disconnect from db.
        :return:
        """
        self.cursor.close()
        self.connection.close()


if __name__ == '__main__':
    database = DataBase('m.sayad', 'localhost', '5432')
    database.connect_db('animal')
